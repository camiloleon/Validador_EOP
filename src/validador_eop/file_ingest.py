from __future__ import annotations

import csv
import io
import json
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


class UnsupportedInputFormatError(ValueError):
    pass


def convert_to_csv_bytes(filename: str, content: bytes) -> bytes:
    extension = Path(filename or "").suffix.lower()

    if extension in {"", ".csv", ".txt"}:
        return content
    if extension == ".json":
        return _json_to_csv(content).encode("utf-8")
    if extension in {".xlsx", ".xlsm", ".xltx", ".xltm"}:
        return _xlsx_to_csv(content).encode("utf-8")
    if extension == ".xls":
        return _xls_to_csv(content).encode("utf-8")

    raise UnsupportedInputFormatError(
        "Formato no soportado. Usa csv, json, xlsx o xls."
    )


def _to_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _safe_header(value: Any, index: int) -> str:
    text = _to_text(value)
    return text if text else f"col_{index}"


def _non_empty_cells(values: list[Any] | tuple[Any, ...]) -> int:
    return sum(1 for value in values if _to_text(value) != "")


def _select_header_index(rows: list[list[Any]]) -> int | None:
    best_index: int | None = None
    best_non_empty = 0
    for index, row in enumerate(rows):
        non_empty = _non_empty_cells(row)
        if non_empty == 0:
            continue
        if non_empty > best_non_empty:
            best_non_empty = non_empty
            best_index = index
    return best_index


def _build_tabular_from_rows(rows: list[list[Any]]) -> tuple[list[str], list[list[str]]]:
    header_index = _select_header_index(rows)
    if header_index is None:
        return [], []

    header_source = rows[header_index]
    headers = [_safe_header(value, index + 1) for index, value in enumerate(header_source)]

    data_rows: list[list[str]] = []
    for row in rows[header_index + 1 :]:
        values = [_to_text(value) for value in row[: len(headers)]]
        if len(values) < len(headers):
            values.extend([""] * (len(headers) - len(values)))
        if any(value != "" for value in values):
            data_rows.append(values)

    return headers, data_rows


def _xlsx_to_csv(content: bytes) -> str:
    workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    best_headers: list[str] = []
    best_rows: list[list[str]] = []
    best_score = -1

    for sheet in workbook.worksheets:
        sheet_rows = [list(row) for row in sheet.iter_rows(values_only=True)]
        headers, data_rows = _build_tabular_from_rows(sheet_rows)
        if not headers:
            continue

        header_non_empty = sum(1 for header in headers if header and not header.startswith("col_"))
        score = (len(data_rows) * 1000) + header_non_empty
        if score > best_score:
            best_score = score
            best_headers = headers
            best_rows = data_rows

    if not best_headers:
        workbook.close()
        raise UnsupportedInputFormatError("El archivo Excel no contiene datos.")

    output = io.StringIO()
    writer = csv.writer(output, delimiter=",", lineterminator="\n")
    writer.writerow(best_headers)

    for row in best_rows:
        writer.writerow(row)

    workbook.close()

    return output.getvalue()


def _xls_to_csv(content: bytes) -> str:
    try:
        import xlrd
    except ModuleNotFoundError as exc:
        raise UnsupportedInputFormatError(
            "Para archivos .xls instala 'xlrd' o convierte a xlsx/csv."
        ) from exc

    workbook = xlrd.open_workbook(file_contents=content)
    if workbook.nsheets == 0:
        raise UnsupportedInputFormatError("El archivo XLS no contiene hojas.")

    best_headers: list[str] = []
    best_rows: list[list[str]] = []
    best_score = -1
    for sheet_idx in range(workbook.nsheets):
        sheet = workbook.sheet_by_index(sheet_idx)
        if sheet.nrows == 0:
            continue

        raw_rows: list[list[Any]] = []
        for row_idx in range(sheet.nrows):
            raw_rows.append([sheet.cell_value(row_idx, col) for col in range(sheet.ncols)])

        headers, data_rows = _build_tabular_from_rows(raw_rows)
        if not headers:
            continue

        header_non_empty = sum(1 for header in headers if header and not header.startswith("col_"))
        score = (len(data_rows) * 1000) + header_non_empty
        if score > best_score:
            best_score = score
            best_headers = headers
            best_rows = data_rows

    if not best_headers:
        raise UnsupportedInputFormatError("El archivo XLS no contiene datos.")

    output = io.StringIO()
    writer = csv.writer(output, delimiter=",", lineterminator="\n")
    writer.writerow(best_headers)

    for values in best_rows:
        writer.writerow(values)

    return output.getvalue()


def _json_to_csv(content: bytes) -> str:
    try:
        parsed = json.loads(content.decode("utf-8-sig", errors="replace"))
    except json.JSONDecodeError as exc:
        raise UnsupportedInputFormatError("JSON inválido.") from exc

    records: list[dict[str, Any]]
    if isinstance(parsed, list):
        records = [item for item in parsed if isinstance(item, dict)]
    elif isinstance(parsed, dict):
        payload = parsed.get("rows") or parsed.get("data") or parsed.get("items")
        if isinstance(payload, list):
            records = [item for item in payload if isinstance(item, dict)]
        else:
            records = [parsed]
    else:
        raise UnsupportedInputFormatError("JSON no compatible. Usa lista de objetos.")

    if not records:
        raise UnsupportedInputFormatError("JSON sin registros.")

    headers: list[str] = []
    for record in records:
        for key in record.keys():
            normalized = str(key)
            if normalized not in headers:
                headers.append(normalized)

    output = io.StringIO()
    writer = csv.DictWriter(output, fieldnames=headers, delimiter=",", lineterminator="\n")
    writer.writeheader()

    for record in records:
        writer.writerow({header: _to_text(record.get(header)) for header in headers})

    return output.getvalue()
