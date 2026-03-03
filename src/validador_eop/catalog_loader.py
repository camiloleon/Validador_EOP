from __future__ import annotations

import csv
from datetime import datetime
from pathlib import Path

import openpyxl

from .normalization import normalize_key


class Catalogs:
    def __init__(
        self,
        roles: set[str],
        areas: set[str],
        trabajos: set[str],
        trabajos_por_area: dict[str, set[str]],
        ciudades: set[str],
        ciudad_to_base: dict[str, str],
        ciudad_to_regional: dict[str, str],
        bases: set[str],
        regionales: set[str],
        companias_nit: set[str],
        existing_tecnicos_ids: set[str] | None = None,
        existing_tecnicos_name_email: set[tuple[str, str]] | None = None,
        existing_tecnicos_name_phone: set[tuple[str, str]] | None = None,
        existing_tecnicos_snapshot_date: str | None = None,
    ) -> None:
        self.roles = roles
        self.areas = areas
        self.trabajos = trabajos
        self.trabajos_por_area = trabajos_por_area
        self.ciudades = ciudades
        self.ciudad_to_base = ciudad_to_base
        self.ciudad_to_regional = ciudad_to_regional
        self.bases = bases
        self.regionales = regionales
        self.companias_nit = companias_nit
        self.existing_tecnicos_ids = existing_tecnicos_ids or set()
        self.existing_tecnicos_name_email = existing_tecnicos_name_email or set()
        self.existing_tecnicos_name_phone = existing_tecnicos_name_phone or set()
        self.existing_tecnicos_snapshot_date = existing_tecnicos_snapshot_date


def _safe_text(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _normalize_phone(value: str) -> str:
    return "".join(ch for ch in value if ch.isdigit())


def _normalize_identifier(value: str) -> str:
    digits = "".join(ch for ch in value if ch.isdigit())
    return digits or normalize_key(value)


def _parse_snapshot_date(value: str) -> datetime | None:
    text = _safe_text(value)
    if not text:
        return None

    for fmt in ("%Y-%m-%d %H:%M", "%Y-%m-%d %H:%M:%S", "%Y-%m-%d"):
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            continue
    return None


def _load_existing_tecnicos(csv_path: Path) -> tuple[set[str], set[tuple[str, str]], set[tuple[str, str]], str | None]:
    existing_ids: set[str] = set()
    existing_name_email: set[tuple[str, str]] = set()
    existing_name_phone: set[tuple[str, str]] = set()
    latest_snapshot: datetime | None = None

    if not csv_path.exists():
        return existing_ids, existing_name_email, existing_name_phone, None

    with csv_path.open("r", encoding="utf-8-sig", newline="") as file:
        reader = csv.DictReader(file)
        field_map = {normalize_key(name).lower(): name for name in (reader.fieldnames or [])}

        id_key = field_map.get("identificacion")
        name_key = field_map.get("nombre_completo")
        email_key = field_map.get("email")
        phone_key = field_map.get("celular")
        date_key = field_map.get("fecha_creacion")

        for row in reader:
            identifier = _normalize_identifier(_safe_text(row.get(id_key))) if id_key else ""
            full_name = normalize_key(_safe_text(row.get(name_key))) if name_key else ""
            email = _safe_text(row.get(email_key)).lower() if email_key else ""
            phone = _normalize_phone(_safe_text(row.get(phone_key))) if phone_key else ""

            if identifier:
                existing_ids.add(identifier)
            if full_name and email:
                existing_name_email.add((full_name, email))
            if full_name and phone:
                existing_name_phone.add((full_name, phone))

            if date_key:
                parsed = _parse_snapshot_date(_safe_text(row.get(date_key)))
                if parsed is not None and (latest_snapshot is None or parsed > latest_snapshot):
                    latest_snapshot = parsed

    snapshot_text = latest_snapshot.strftime("%Y-%m-%d %H:%M") if latest_snapshot else None
    return existing_ids, existing_name_email, existing_name_phone, snapshot_text


def load_catalogs_from_excel(excel_path: str | Path) -> Catalogs:
    excel_path = Path(excel_path)
    workbook = openpyxl.load_workbook(excel_path, data_only=True)

    roles: set[str] = set()
    areas: set[str] = set()
    trabajos: set[str] = set()
    trabajos_por_area: dict[str, set[str]] = {}
    ciudades: set[str] = set()
    ciudad_to_base: dict[str, str] = {}
    ciudad_to_regional: dict[str, str] = {}
    bases: set[str] = set()
    regionales: set[str] = set()
    companias_nit: set[str] = set()
    existing_tecnicos_ids: set[str] = set()
    existing_tecnicos_name_email: set[tuple[str, str]] = set()
    existing_tecnicos_name_phone: set[tuple[str, str]] = set()
    existing_tecnicos_snapshot_date: str | None = None

    tecnicos_export_path = excel_path.parent / "tecnicos_exportados.csv"
    (
        existing_tecnicos_ids,
        existing_tecnicos_name_email,
        existing_tecnicos_name_phone,
        existing_tecnicos_snapshot_date,
    ) = _load_existing_tecnicos(tecnicos_export_path)

    if "Roles" in workbook.sheetnames:
        sheet = workbook["Roles"]
        for row in sheet.iter_rows(min_row=2, values_only=True):
            role = normalize_key(_safe_text(row[1] if len(row) > 1 else None))
            if role:
                roles.add(role)

    if "Areas" in workbook.sheetnames:
        sheet = workbook["Areas"]
        for row in sheet.iter_rows(min_row=2, values_only=True):
            area = normalize_key(_safe_text(row[1] if len(row) > 1 else None))
            if area:
                areas.add(area)

    if "Tipo Trabajos" in workbook.sheetnames:
        sheet = workbook["Tipo Trabajos"]
        for row in sheet.iter_rows(min_row=2, values_only=True):
            trabajo = normalize_key(_safe_text(row[1] if len(row) > 1 else None))
            area = normalize_key(_safe_text(row[2] if len(row) > 2 else None))
            if trabajo:
                trabajos.add(trabajo)
                if area:
                    trabajos_por_area.setdefault(area, set()).add(trabajo)

    if "Ciudades" in workbook.sheetnames:
        sheet = workbook["Ciudades"]
        for row in sheet.iter_rows(min_row=2, values_only=True):
            ciudad = normalize_key(_safe_text(row[1] if len(row) > 1 else None))
            base = normalize_key(_safe_text(row[3] if len(row) > 3 else None))
            regional = normalize_key(_safe_text(row[4] if len(row) > 4 else None))
            if ciudad:
                ciudades.add(ciudad)
                if base:
                    ciudad_to_base[ciudad] = base
                if regional:
                    ciudad_to_regional[ciudad] = regional

    if "Bases Operativas" in workbook.sheetnames:
        sheet = workbook["Bases Operativas"]
        for row in sheet.iter_rows(min_row=2, values_only=True):
            base = normalize_key(_safe_text(row[1] if len(row) > 1 else None))
            regional = normalize_key(_safe_text(row[4] if len(row) > 4 else None))
            if base:
                bases.add(base)
            if regional:
                regionales.add(regional)

    if "Regionales" in workbook.sheetnames:
        sheet = workbook["Regionales"]
        for row in sheet.iter_rows(min_row=2, values_only=True):
            regional = normalize_key(_safe_text(row[1] if len(row) > 1 else None))
            if regional:
                regionales.add(regional)

    if "Compa�ias" in workbook.sheetnames:
        sheet = workbook["Compa�ias"]
    elif "Compañias" in workbook.sheetnames:
        sheet = workbook["Compañias"]
    else:
        sheet = None

    if sheet is not None:
        for row in sheet.iter_rows(min_row=2, values_only=True):
            nit = normalize_key(_safe_text(row[0] if len(row) > 0 else None))
            if nit:
                companias_nit.add(nit)

    return Catalogs(
        roles=roles,
        areas=areas,
        trabajos=trabajos,
        trabajos_por_area=trabajos_por_area,
        ciudades=ciudades,
        ciudad_to_base=ciudad_to_base,
        ciudad_to_regional=ciudad_to_regional,
        bases=bases,
        regionales=regionales,
        companias_nit=companias_nit,
        existing_tecnicos_ids=existing_tecnicos_ids,
        existing_tecnicos_name_email=existing_tecnicos_name_email,
        existing_tecnicos_name_phone=existing_tecnicos_name_phone,
        existing_tecnicos_snapshot_date=existing_tecnicos_snapshot_date,
    )
