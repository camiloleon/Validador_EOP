from __future__ import annotations

import json
from io import BytesIO

from openpyxl import Workbook

from src.validador_eop.file_ingest import UnsupportedInputFormatError, convert_to_csv_bytes


def test_csv_passthrough() -> None:
    raw = b"a,b\n1,2\n"
    out = convert_to_csv_bytes("entrada.csv", raw)
    assert out == raw


def test_json_to_csv() -> None:
    payload = json.dumps(
        [
            {"identificacion": "100", "correo": "uno@mail.com"},
            {"identificacion": "200", "correo": "dos@mail.com"},
        ]
    ).encode("utf-8")

    out = convert_to_csv_bytes("entrada.json", payload).decode("utf-8")
    assert "identificacion,correo" in out
    assert "100,uno@mail.com" in out


def test_xlsx_to_csv() -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["identificacion", "correo"])
    sheet.append([123, "uno@mail.com"])

    stream = BytesIO()
    workbook.save(stream)

    out = convert_to_csv_bytes("entrada.xlsx", stream.getvalue()).decode("utf-8")
    assert "identificacion,correo" in out
    assert "123,uno@mail.com" in out


def test_xlsx_prefers_sheet_with_real_data() -> None:
    workbook = Workbook()
    portada = workbook.active
    portada.title = "Portada"
    portada["A1"] = "Instructivo"

    usuarios = workbook.create_sheet("Usuarios")
    usuarios.append(["nombre completo", "email", "identificacion", "rol de usuario", "regional", "compania (nit)"])
    usuarios.append(["Ana Lopez", "ana@mail.com", "123", "TECNICO", "R2", "900123456-7"])
    usuarios.append(["Luis Perez", "luis@mail.com", "456", "OWNER", "R1", "900123456-7"])

    stream = BytesIO()
    workbook.save(stream)

    out = convert_to_csv_bytes("usuarios.xlsx", stream.getvalue()).decode("utf-8")
    assert "nombre completo,email,identificacion,rol de usuario,regional,compania (nit)" in out
    assert "Ana Lopez,ana@mail.com,123,TECNICO,R2,900123456-7" in out
    assert "Luis Perez,luis@mail.com,456,OWNER,R1,900123456-7" in out


def test_xlsx_detects_header_not_on_first_row() -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Usuarios"
    sheet.append(["Archivo generado por sistema"]) 
    sheet.append([""])
    sheet.append(["nombre completo", "email", "identificacion", "rol de usuario", "regional", "compania (nit)"])
    sheet.append(["Maria R", "maria@mail.com", "111", "TECNICO", "R2", "900123456-7"])

    stream = BytesIO()
    workbook.save(stream)

    out = convert_to_csv_bytes("usuarios.xlsx", stream.getvalue()).decode("utf-8")
    assert "nombre completo,email,identificacion,rol de usuario,regional,compania (nit)" in out
    assert "Maria R,maria@mail.com,111,TECNICO,R2,900123456-7" in out


def test_unsupported_extension() -> None:
    try:
        convert_to_csv_bytes("entrada.xml", b"<x></x>")
    except UnsupportedInputFormatError as exc:
        assert "Formato no soportado" in str(exc)
    else:
        raise AssertionError("Se esperaba UnsupportedInputFormatError")
